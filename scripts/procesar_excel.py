import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


def print_json_message(message, success=False, warning=False, error=False):
    msg_type = "info"
    if success:
        msg_type = "success"
    elif warning:
        msg_type = "warning"
    elif error:
        msg_type = "error"
    print(json.dumps({"type": msg_type, "message": message}), file=sys.stderr)


def parse_args():
    if len(sys.argv) < 2:
        print_json_message("No se recibió el archivo de entrada.", error=True)
        sys.exit(1)

    input_path = sys.argv[1]
    lanes_ns_arg = sys.argv[2] if len(sys.argv) > 2 else ""
    lanes_sn_arg = sys.argv[3] if len(sys.argv) > 3 else ""
    intervals_arg = sys.argv[4] if len(sys.argv) > 4 else ""

    def parse_lanes(lanes_str):
        if lanes_str:
            try:
                return list(dict.fromkeys(int(x.strip()) for x in lanes_str.split(",") if x.strip()))
            except Exception:
                print_json_message(f"Formato inválido de carriles: {lanes_str}", warning=True)
                return []
        return []

    lanes_ns = parse_lanes(lanes_ns_arg)
    lanes_sn = parse_lanes(lanes_sn_arg)

    if not lanes_ns and not lanes_sn:
        print_json_message("No se especificaron carriles. Se procesarán todos.", warning=True)

    intervals = []
    if intervals_arg:
        for p in [p.strip() for p in intervals_arg.split(",") if p.strip()]:
            if "-" in p:
                a, b = p.split("-", 1)
                intervals.append((a.strip(), b.strip()))
    else:
        print_json_message("No se especificaron intervalos. Se tomará el día completo.", warning=True)

    return input_path, lanes_ns, lanes_sn, intervals


def read_file_safely(input_path):
    try:
        with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
            sample = f.read(2048)
            f.seek(0)
            sep = ";" if ";" in sample and "," not in sample else ","
            if "," in sample or ";" in sample or "\t" in sample:
                print_json_message("Archivo CSV detectado, leyendo con pandas.", success=True)
                df = pd.read_csv(f, sep=sep)
            else:
                print_json_message("Archivo Excel detectado.", success=True)
                df = pd.read_excel(input_path, engine="openpyxl")
    except Exception:
        try:
            print_json_message("Intentando leer como Excel.", warning=True)
            df = pd.read_excel(input_path, engine="openpyxl")
        except Exception as e2:
            print_json_message(f"Error leyendo archivo: {str(e2)}", error=True)
            print("[]")
            sys.exit(0)
    return df


def procesar_direccion(df, carriles, intervalos, nombre):
    resultados = []
    if carriles:
        df_dir = df[df["Lane"].isin(carriles)]
    else:
        df_dir = df.copy()

    if df_dir.empty:
        print_json_message(f"No se encontraron datos para {nombre}.", warning=True)
        return pd.DataFrame()

    if "#vehicles" not in df_dir.columns:
        total_por_fila = pd.Series(0, index=df_dir.index, dtype="float64")
        for col in df_dir.columns:
            if "vehicle" in str(col).lower():
                total_por_fila = pd.to_numeric(df_dir[col], errors="coerce").fillna(0)
                break
        df_dir = df_dir.copy()
        df_dir["#vehicles"] = total_por_fila

    for fecha, grupo in df_dir.groupby("Fecha"):
        grupo_filtrado = grupo.copy()

        if intervalos:
            mascara = pd.Series(False, index=grupo_filtrado.index)
            for inicio, fin in intervalos:
                inicio_hora = pd.to_datetime(f"{fecha} {inicio}")
                fin_hora = pd.to_datetime(f"{fecha} {fin}")
                mascara = mascara | ((grupo_filtrado["Time"] >= inicio_hora) & (grupo_filtrado["Time"] < fin_hora))
            grupo_filtrado = grupo_filtrado[mascara]

        if grupo_filtrado.empty:
            continue

        agg = (
            grupo_filtrado.groupby("HoraMinuto", as_index=False)["#vehicles"]
            .sum()
            .sort_values("HoraMinuto")
        )

        for _, fila in agg.iterrows():
            resultados.append({
                "Dirección": nombre,
                "Fecha": str(fecha),
                "Intervalo": fila["HoraMinuto"],
                "Carriles": ",".join(map(str, carriles)) if carriles else "all",
                "Total_vehiculos": int(fila["#vehicles"]),
            })

    print_json_message(f"Procesados {len(resultados)} registros para {nombre}.", success=True)
    return pd.DataFrame(resultados)


def main():
    input_path, lanes_ns, lanes_sn, intervalos = parse_args()
    df = read_file_safely(input_path)
    df.columns = df.columns.str.strip()

    rename_map = {}
    if "Utc" in df.columns:
        rename_map["Utc"] = "Time"
    if "ZoneId" in df.columns:
        rename_map["ZoneId"] = "Lane"
    if "NumVeh" in df.columns:
        rename_map["NumVeh"] = "#vehicles"
    df = df.rename(columns=rename_map)

    if "Time" not in df.columns:
        print_json_message("Falta la columna 'Time' en el archivo.", error=True)
        print("[]")
        sys.exit(0)

    df["Time"] = pd.to_datetime(df["Time"], errors="coerce", dayfirst=True)
    if isinstance(df["Time"].dtype, pd.DatetimeTZDtype):
        df["Time"] = df["Time"].dt.tz_convert(None)
    df = df.dropna(subset=["Time"])
    df["Fecha"] = df["Time"].dt.date

    if "Lane" in df.columns:
        df["Lane"] = pd.to_numeric(df["Lane"], errors="coerce")
        df = df.dropna(subset=["Lane"])
        df["Lane"] = df["Lane"].astype(int)

    if "#vehicles" in df.columns:
        df["#vehicles"] = pd.to_numeric(df["#vehicles"], errors="coerce").fillna(0)

    df["HoraMinuto"] = df["Time"].dt.strftime("%H:%M")

    if "#vehicles" not in df.columns:
        total_por_fila = pd.Series(0, index=df.index, dtype="float64")
        for col in df.columns:
            if "vehicle" in str(col).lower():
                total_por_fila = pd.to_numeric(df[col], errors="coerce").fillna(0)
                break
        df["#vehicles"] = total_por_fila

    before_dedup = len(df)
    df = (
        df.sort_values("#vehicles", ascending=False)
        .drop_duplicates(subset=["Lane", "Fecha", "HoraMinuto"], keep="first")
        .sort_values("Time")
    )
    print_json_message(
        f"Deduplicación por Lane+Fecha+Hora aplicada: {before_dedup - len(df)} duplicados eliminados.",
        success=True,
    )

    if df.empty:
        print_json_message("El archivo no contiene datos válidos.", error=True)
        print("[]")
        sys.exit(0)

    df_ns = procesar_direccion(df, lanes_ns, intervalos, "Norte → Sur")
    df_sn = procesar_direccion(df, lanes_sn, intervalos, "Sur → Norte")

    resultados = pd.concat([df_ns, df_sn], ignore_index=True)
    if resultados.empty:
        print_json_message("No se generaron resultados, revise filtros y carriles.", warning=True)
    else:
        print_json_message("Procesamiento completado exitosamente.", success=True)

    if not resultados.empty and "Fecha" in resultados.columns:
        resultados["Fecha"] = pd.to_datetime(resultados["Fecha"]).dt.strftime("%d/%m/%Y")

    salida_json = resultados.to_dict(orient="records")

    if "--write-xlsx" in sys.argv:
        with pd.ExcelWriter("resumen_intervalos.xlsx", engine="openpyxl") as writer:
            df_ns.to_excel(writer, index=False, sheet_name="Norte_Sur")
            df_sn.to_excel(writer, index=False, sheet_name="Sur_Norte")

        wb = load_workbook("resumen_intervalos.xlsx")
        thin_border = Border(bottom=Side(style="thin", color="000000"))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx in range(2, ws.max_row + 1):
                if (row_idx - 1) % 3 == 0:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).border = thin_border
        wb.save("resumen_intervalos.xlsx")
        print_json_message("Archivo Excel generado correctamente.", success=True)

    print(json.dumps(salida_json, ensure_ascii=False))


if __name__ == "__main__":
    main()
